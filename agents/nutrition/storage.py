"""
storage.py — Persistencia del registro nutricional.

Dos ficheros independientes:
  comidas.xlsx  → registro diario con macros acumulados (formato visual)
  alimentos.csv → historial de alimentos por fecha (para generar menús)
"""

import csv
import io
import os
import logging
from datetime import datetime, timedelta
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

logger = logging.getLogger(__name__)

# ── Configuración ─────────────────────────────────────────────────────────────

EXCEL_PATH = Path(os.getenv("NUTRITION_EXCEL_PATH", "/data/comidas.xlsx"))
BASE_PATH  = Path(os.getenv("BASE_NUTRICIONAL_PATH", "/data/base_nutricional.xlsx"))
CSV_PATH   = Path(os.getenv("NUTRITION_CSV_PATH",   "/data/alimentos.csv"))

COLUMNAS = [
    "Fecha", "Descripcion",
    "Calorias (kcal)", "Proteinas (g)", "Carbohidratos (g)",
    "Grasas (g)", "Azucar (g)", "Fibra (g)",
]
_COL = {nombre: i + 1 for i, nombre in enumerate(COLUMNAS)}

_HEADER_BG = "6B8CAE"
_HEADER_FG = "FFFFFF"
_ROW_BG    = ["EAF4FB", "F5FBFE"]
_BORDER_C  = "BBCDD8"
_NUM_FMT   = "#,##0.0"

_CSV_CAMPOS = ["fecha", "alimento"]


# ── Estilos ───────────────────────────────────────────────────────────────────

def _make_border() -> Border:
    side = Side(style="thin", color=_BORDER_C)
    return Border(left=side, right=side, top=side, bottom=side)


def _style(cell, *, bold=False, fg="000000", bg: str | None = None,
           align="center", size=10) -> None:
    cell.font      = Font(name="Arial", bold=bold, size=size, color=fg)
    cell.alignment = Alignment(
        horizontal=align, vertical="center",
        indent=1 if align == "left" else 0,
    )
    cell.border = _make_border()
    if bg:
        cell.fill = PatternFill("solid", start_color=bg)


# ── Crear / cargar Excel ──────────────────────────────────────────────────────

def _crear_excel() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Comidas"

    last_col = chr(64 + len(COLUMNAS))
    ws.merge_cells(f"A1:{last_col}1")
    _style(ws["A1"], bold=True, fg=_HEADER_FG, bg=_HEADER_BG, size=13)
    ws["A1"].value = "Registro Nutricional"
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 6

    for i, nombre in enumerate(COLUMNAS, 1):
        cell = ws.cell(row=3, column=i, value=nombre)
        _style(cell, bold=True, fg=_HEADER_FG, bg=_HEADER_BG, size=11)

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 45
    for letra in "CDEFGH":
        ws.column_dimensions[letra].width = 18

    EXCEL_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(EXCEL_PATH)
    logger.info("Excel creado en %s", EXCEL_PATH)
    return wb


def _cargar_excel() -> Workbook:
    if not EXCEL_PATH.exists():
        return _crear_excel()
    return load_workbook(EXCEL_PATH)


# ── Crear / cargar CSV ────────────────────────────────────────────────────────

def _inicializar_csv() -> None:
    """Crea el CSV con cabeceras si no existe. Se llama solo cuando hace falta."""
    if not CSV_PATH.exists():
        CSV_PATH.parent.mkdir(parents=True, exist_ok=True)
        with open(CSV_PATH, "w", newline="", encoding="utf-8") as f:
            csv.writer(f).writerow(_CSV_CAMPOS)
        logger.info("CSV de alimentos creado en %s", CSV_PATH)


# ── Helpers internos ──────────────────────────────────────────────────────────

def _fila_de_fecha(ws, fecha_str: str) -> int | None:
    for row in ws.iter_rows(min_row=4, max_col=1):
        if str(row[0].value or "").strip() == fecha_str:
            return row[0].row
    return None


def _siguiente_fila_libre(ws) -> int:
    max_row = 3
    for row in ws.iter_rows(min_row=4, max_col=1):
        if row[0].value is not None:
            max_row = row[0].row
    return max_row + 1


def _escribir_fila(ws, row_num: int, fecha_str: str,
                   datos: dict, es_nueva: bool) -> None:
    bg = _ROW_BG[row_num % 2]

    def celda(col_name: str, valor, fmt: str | None = None,
              align: str = "center") -> None:
        c = ws.cell(row=row_num, column=_COL[col_name], value=valor)
        _style(c, bg=bg, align=align)
        if fmt:
            c.number_format = fmt

    macros = [
        ("Calorias (kcal)",   "calorias"),
        ("Proteinas (g)",     "proteinas"),
        ("Carbohidratos (g)", "carbohidratos"),
        ("Grasas (g)",        "grasas"),
        ("Azucar (g)",        "azucar"),
        ("Fibra (g)",         "fibra"),
    ]

    if es_nueva:
        celda("Fecha",       fecha_str,                   align="center")
        celda("Descripcion", datos["descripcion_comida"], align="left")
        for col_name, key in macros:
            celda(col_name, float(datos["totales"][key]), _NUM_FMT)
    else:
        for col_name, key in macros:
            c = ws.cell(row=row_num, column=_COL[col_name])
            c.value = round(float(c.value or 0) + datos["totales"][key], 1)
            _style(c, bg=bg)
            c.number_format = _NUM_FMT
        c = ws.cell(row=row_num, column=_COL["Descripcion"])
        c.value = f"{c.value or ''} + {datos['descripcion_comida']}"
        _style(c, bg=bg, align="left")

    ws.row_dimensions[row_num].height = 20


# ── API pública ───────────────────────────────────────────────────────────────

def guardar_comida(datos: dict, fecha: datetime) -> tuple[bool, bool]:
    """
    Guarda o acumula macros en el Excel.
    Guarda los nombres de alimentos en el CSV.

    Returns:
        (ok, es_nueva_fila)
    """
    # Validar alimentos antes de tocar ningún fichero
    alimentos = datos.get("alimentos")
    if not isinstance(alimentos, list):
        logger.error("Campo 'alimentos' no es una lista: %r", alimentos)
        return False, False

    try:
        fecha_str = fecha.strftime("%d/%m/%Y")

        # ── Excel: acumular totales del día ───────────────────────────────────
        wb       = _cargar_excel()
        ws       = wb["Comidas"]
        fila     = _fila_de_fecha(ws, fecha_str)
        es_nueva = fila is None
        row_num  = _siguiente_fila_libre(ws) if es_nueva else fila

        _escribir_fila(ws, row_num, fecha_str, datos, es_nueva)
        wb.save(EXCEL_PATH)

        # ── CSV: una línea por alimento ───────────────────────────────────────
        _inicializar_csv()
        with open(CSV_PATH, "a", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            for a in alimentos:
                nombre = str(a.get("nombre", "")).strip()
                if nombre:
                    writer.writerow([fecha_str, nombre])

        return True, es_nueva

    except Exception:
        logger.exception("Error guardando comida")
        return False, False


def leer_registros(dias: int = 7) -> list[dict]:
    """
    Lee los registros de los últimos N días desde el Excel.
    dias=0 → todos los registros.
    """
    if not EXCEL_PATH.exists():
        return []
    try:
        df = pd.read_excel(EXCEL_PATH, sheet_name="Comidas", header=2)
        df = df.dropna(subset=[df.columns[0]])

        limite = (datetime.now() - timedelta(days=dias)) if dias > 0 else None
        registros = []

        for _, row in df.iterrows():
            try:
                fecha_dt = datetime.strptime(str(row.iloc[0]).strip(), "%d/%m/%Y")
            except ValueError:
                continue
            if limite and fecha_dt < limite:
                continue
            registros.append({
                "fecha":         str(row.iloc[0]).strip(),
                "descripcion":   str(row.iloc[1] or ""),
                "calorias":      float(row.iloc[2] or 0),
                "proteinas":     float(row.iloc[3] or 0),
                "carbohidratos": float(row.iloc[4] or 0),
                "grasas":        float(row.iloc[5] or 0),
                "azucar":        float(row.iloc[6] or 0),
                "fibra":         float(row.iloc[7] or 0),
            })

        return sorted(registros,
                      key=lambda r: datetime.strptime(r["fecha"], "%d/%m/%Y"))
    except Exception:
        logger.exception("Error leyendo registros del Excel")
        return []


def leer_csv_alimentos(dias: int = 90) -> str:
    """
    Lee el CSV de alimentos de los últimos N días.
    Devuelve el contenido como string para pasárselo directamente a Claude.

    dias=90 → últimos 3 meses, suficiente para detectar hábitos sin exceder contexto.
    """
    if not CSV_PATH.exists():
        return ""
    try:
        limite = datetime.now() - timedelta(days=dias)
        lineas = []

        with open(CSV_PATH, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                try:
                    fecha_dt = datetime.strptime(row["fecha"].strip(), "%d/%m/%Y")
                except (ValueError, KeyError):
                    continue
                if fecha_dt >= limite:
                    lineas.append(row)

        if not lineas:
            return ""

        buf = io.StringIO()
        writer = csv.DictWriter(buf, fieldnames=_CSV_CAMPOS)
        writer.writeheader()
        writer.writerows(lineas)
        return buf.getvalue()

    except Exception:
        logger.exception("Error leyendo CSV de alimentos")
        return ""


def borrar_dia(fecha_str: str) -> bool:
    """
    Elimina la fila de fecha_str en el Excel
    y todas las líneas de esa fecha en el CSV.
    """
    try:
        wb  = _cargar_excel()
        ws  = wb["Comidas"]
        fila = _fila_de_fecha(ws, fecha_str)
        if fila is None:
            return False
        ws.delete_rows(fila)
        wb.save(EXCEL_PATH)

        # Reescribir CSV sin las líneas de esa fecha
        if CSV_PATH.exists():
            lineas_restantes = []
            with open(CSV_PATH, "r", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    if row.get("fecha", "").strip() != fecha_str:
                        lineas_restantes.append(row)
            with open(CSV_PATH, "w", newline="", encoding="utf-8") as f:
                writer = csv.DictWriter(f, fieldnames=_CSV_CAMPOS)
                writer.writeheader()
                writer.writerows(lineas_restantes)

        return True
    except Exception:
        logger.exception("Error borrando fila %s", fecha_str)
        return False


def borrar_todo() -> bool:
    """Elimina el Excel y el CSV, y los recrea vacíos."""
    try:
        if EXCEL_PATH.exists():
            EXCEL_PATH.unlink()
        if CSV_PATH.exists():
            CSV_PATH.unlink()
        _crear_excel()
        _inicializar_csv()
        return True
    except Exception:
        logger.exception("Error reseteando datos")
        return False


def cargar_base_nutricional() -> dict:
    """
    Carga la base nutricional por alimento desde base_nutricional.xlsx.
    Devuelve dict vacío si no existe o hay error.
    """
    if not BASE_PATH.exists():
        return {}
    try:
        df = pd.read_excel(BASE_PATH)
        df.columns = [str(c).strip().lower() for c in df.columns]
        base = {}
        for _, row in df.iterrows():
            nombre = str(row.get("alimento", row.get("nombre", ""))).strip().lower()
            if not nombre:
                continue
            base[nombre] = {
                "calorias":      float(row.get("calorias",      row.get("kcal",     0)) or 0),
                "proteinas":     float(row.get("proteinas",     row.get("proteina", 0)) or 0),
                "carbohidratos": float(row.get("carbohidratos", row.get("hidratos", 0)) or 0),
                "grasas":        float(row.get("grasas",        row.get("grasa",    0)) or 0),
                "azucar":        float(row.get("azucar",  0) or 0),
                "fibra":         float(row.get("fibra",   0) or 0),
            }
        return base
    except Exception:
        logger.warning("No se pudo cargar base_nutricional.xlsx")
        return {}