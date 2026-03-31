"""
agente_compra.py  — Lista de la compra para Lisa Bot
Versión corregida y refactorizada.

Cambios respecto al original:
  - leer_items(): lectura defensiva, nunca devuelve basura ni borra datos
  - guardar_excel(): escritura atómica (tmp → rename) para evitar corrupción
  - parsear_mensaje(): prompt mejorado, manejo robusto de JSON
  - Filtros por tienda + categoría simultáneos (ver_filtro_combinado)
  - Confirmación cuando el borrado afecta a más de 5 artículos
  - Todas las funciones de vista revisadas y consistentes
"""

import os
import json
import logging
import tempfile
from collections import defaultdict
from dotenv import load_dotenv
from anthropic import Anthropic
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from telegram import InlineKeyboardButton, InlineKeyboardMarkup

load_dotenv()

logger = logging.getLogger(__name__)

EXCEL_PATH = os.getenv("EXCEL_PATH", "/data/lista_compra.xlsx")
claude = Anthropic(api_key=os.getenv("ANTHROPIC_API_KEY"))

# ─────────────────────────────────────────
# CONSTANTES
# ─────────────────────────────────────────

# Número mínimo de artículos a borrar para pedir confirmación
UMBRAL_CONFIRMACION = 5

CATEGORIAS: dict[str, str] = {
    "alimentación":        "🥦",
    "higiene personal":    "🧴",
    "limpieza hogar":      "🧹",
    "farmacia y salud":    "💊",
    "tecnología":          "💻",
    "electrodomésticos":   "🔌",
    "mobiliario":          "🛋️",
    "textil y ropa":       "👕",
    "papelería y oficina": "📎",
    "otros":               "📦",
}

CATEGORIAS_VALIDAS = set(CATEGORIAS.keys())

ACCIONES_QUE_BORRAN = {"eliminar", "limpiar_tienda", "limpiar_categoria"}


def emoji_categoria(cat: str) -> str:
    return CATEGORIAS.get(cat.lower().strip(), "📦")


# ─────────────────────────────────────────
# ESTILOS EXCEL
# ─────────────────────────────────────────

C_HEADER_BG  = "6B8CAE"
C_HEADER_FG  = "FFFFFF"
C_SECTION_BG = "A8C4D4"
C_SECTION_FG = "2C3E50"
C_ROW1_BG    = "EAF4FB"
C_ROW2_BG    = "F5FBFE"
C_TOTAL_BG   = "7DADA0"
C_TOTAL_FG   = "FFFFFF"
BORDER_COLOR = "BBCDD8"

_thin   = Side(style="thin", color=BORDER_COLOR)
_border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _apply(cell, bold=False, fg="000000", bg=None, align="center", size=10, italic=False):
    cell.font      = Font(name="Arial", bold=bold, size=size, color=fg, italic=italic)
    if bg:
        cell.fill  = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(
        horizontal=align, vertical="center",
        indent=1 if align == "left" else 0,
    )
    cell.border = _border


# ─────────────────────────────────────────
# EXCEL: LEER
# ─────────────────────────────────────────

# Valores que aparecen en el Excel como cabeceras / totales y NO son artículos
_FILAS_IGNORAR = {
    "total", "lista de la compra", "producto", "tienda",
    "cantidad", "prioridad", "categoría", "categoria",
}


def leer_items() -> list[dict]:
    """
    Lee el Excel y devuelve la lista de artículos.
    Es defensiva: si el fichero no existe o está corrupto devuelve [].
    Nunca modifica el fichero.
    """
    if not os.path.exists(EXCEL_PATH):
        logger.debug("leer_items: fichero no encontrado, devolviendo lista vacía")
        return []

    try:
        wb = load_workbook(EXCEL_PATH, data_only=True, read_only=True)
        ws = wb.active
        items: list[dict] = []

        for row in ws.iter_rows(min_row=4, values_only=True):
            # Fila vacía o sin producto → ignorar
            if not row or not row[0]:
                continue

            producto = str(row[0]).strip()

            # Cabeceras, secciones y totales → ignorar
            if producto.lower() in _FILAS_IGNORAR:
                continue
            if producto.lower().startswith("total:"):
                continue
            # Filas de categoría (emoji al principio) o tienda (📍) → ignorar
            if producto.startswith("🥦") or producto.startswith("📍") or \
               any(producto.startswith(e) for e in CATEGORIAS.values()):
                continue

            tienda_raw = str(row[1]).strip() if len(row) > 1 and row[1] else ""

            # Si el campo tienda es una cabecera de sección, ignorar la fila
            if tienda_raw.lower() in _FILAS_IGNORAR:
                continue

            # Si no tiene tienda asignada, guardar como "sin tienda"
            tienda = tienda_raw if tienda_raw else "sin tienda"

            cantidad  = str(row[2]).strip() if len(row) > 2 and row[2] else "1"
            prioridad = str(row[3]).strip().lower() if len(row) > 3 and row[3] else "normal"
            categoria = str(row[4]).strip().lower() if len(row) > 4 and row[4] else "otros"

            # Sanear prioridad
            if prioridad not in ("urgente", "normal"):
                prioridad = "normal"

            # Sanear categoría
            if categoria not in CATEGORIAS_VALIDAS:
                categoria = "otros"

            items.append({
                "producto":  producto,
                "tienda":    tienda.lower(),
                "cantidad":  cantidad,
                "prioridad": prioridad,
                "categoria": categoria,
            })

        wb.close()
        return items

    except Exception as exc:
        logger.error(f"leer_items: error leyendo Excel: {exc}")
        return []


# ─────────────────────────────────────────
# EXCEL: GUARDAR (escritura atómica)
# ─────────────────────────────────────────

def guardar_excel(items: list[dict]) -> None:
    """
    Guarda la lista en Excel usando escritura atómica:
    escribe en un fichero temporal y luego lo mueve al destino.
    Así, si falla a mitad, el fichero original queda intacto.
    """
    os.makedirs(os.path.dirname(os.path.abspath(EXCEL_PATH)), exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Lista Compra"

    # Anchos de columna
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 20

    # Fila 1: título
    ws.merge_cells("A1:E1")
    ws["A1"] = "Lista de la Compra"
    _apply(ws["A1"], bold=True, fg=C_HEADER_FG, bg=C_HEADER_BG, size=13)
    ws.row_dimensions[1].height = 28

    # Fila 2: separador visual
    ws.row_dimensions[2].height = 6

    # Fila 3: cabeceras
    for col, nombre in enumerate(
        ["Producto", "Tienda", "Cantidad", "Prioridad", "Categoría"], start=1
    ):
        cell = ws.cell(row=3, column=col, value=nombre.upper())
        _apply(cell, bold=True, fg=C_HEADER_FG, bg=C_HEADER_BG, size=11)
    ws.row_dimensions[3].height = 20

    # Agrupar por categoría → tienda
    por_categoria: dict[str, dict[str, list[dict]]] = defaultdict(lambda: defaultdict(list))
    for item in items:
        por_categoria[item["categoria"]][item["tienda"]].append(item)

    row_num = 4
    for categoria in sorted(por_categoria.keys()):
        # Sección de categoría
        ws.merge_cells(f"A{row_num}:E{row_num}")
        ws[f"A{row_num}"] = f"{emoji_categoria(categoria)}  {categoria.upper()}"
        _apply(ws[f"A{row_num}"], bold=True, fg=C_HEADER_FG, bg=C_HEADER_BG, align="left", size=11)
        ws.row_dimensions[row_num].height = 20
        row_num += 1

        for tienda in sorted(por_categoria[categoria].keys()):
            # Sub-sección de tienda
            ws.merge_cells(f"A{row_num}:E{row_num}")
            ws[f"A{row_num}"] = f"    📍 {tienda.upper()}"
            _apply(ws[f"A{row_num}"], bold=True, fg=C_SECTION_FG, bg=C_SECTION_BG, align="left", size=10)
            ws.row_dimensions[row_num].height = 18
            row_num += 1

            for i, item in enumerate(por_categoria[categoria][tienda]):
                bg = C_ROW1_BG if i % 2 == 0 else C_ROW2_BG
                for col, val in enumerate(
                    [item["producto"], item["tienda"], item["cantidad"],
                     item["prioridad"], item["categoria"]],
                    start=1,
                ):
                    cell = ws.cell(row=row_num, column=col, value=val)
                    _apply(cell, fg="000000", bg=bg, align="left" if col == 1 else "center")
                ws.row_dimensions[row_num].height = 16
                row_num += 1

    # Fila de total
    ws.merge_cells(f"A{row_num}:D{row_num}")
    ws[f"A{row_num}"] = f"TOTAL: {len(items)} productos"
    _apply(ws[f"A{row_num}"], bold=True, fg=C_TOTAL_FG, bg=C_TOTAL_BG, size=11)
    _apply(ws.cell(row=row_num, column=5, value=""), bg=C_TOTAL_BG)
    ws.row_dimensions[row_num].height = 20

    # Escritura atómica
    dir_destino = os.path.dirname(os.path.abspath(EXCEL_PATH))
    fd, tmp_path = tempfile.mkstemp(dir=dir_destino, suffix=".xlsx")
    try:
        os.close(fd)
        wb.save(tmp_path)
        os.replace(tmp_path, EXCEL_PATH)
    except Exception:
        os.unlink(tmp_path)
        raise


# ─────────────────────────────────────────
# CLAUDE: PARSEAR MENSAJE
# ─────────────────────────────────────────

_SYSTEM_PARSER = """Eres un parser de listas de la compra. Devuelve ÚNICAMENTE JSON válido,
sin backticks, sin texto previo ni posterior, sin comentarios.

─── ACCIONES ───────────────────────────────────────────────────────────────
añadir            → añadir uno o varios productos
eliminar          → borrar productos concretos ya comprados
limpiar_tienda    → borrar TODOS los productos de una tienda
limpiar_categoria → borrar TODOS los productos de una categoría
actualizar        → cambiar tienda/cantidad/prioridad/categoría de un producto
ver_todo          → ver la lista completa
ver_urgentes      → ver sólo los urgentes
ver_categoria     → ver una categoría concreta
ver_categorias    → ver varias categorías a la vez
ver_tienda        → ver una tienda concreta
ver_filtro        → filtrar combinando tiendas y/o categorías a la vez

─── CATEGORÍAS VÁLIDAS (exactas, en minúsculas) ────────────────────────────
alimentación | higiene personal | limpieza hogar | farmacia y salud |
tecnología | electrodomésticos | mobiliario | textil y ropa |
papelería y oficina | otros

─── REGLAS DE CATEGORIZACIÓN ───────────────────────────────────────────────
alimentación      → comida, bebida, ingredientes, snacks, especias, chocolate
higiene personal  → cuidado corporal, cosmética, dental, champú, maquillaje, perfume, bastoncillos, mascarilla facial
limpieza hogar    → detergentes, bayetas, fregonas, lavavajillas, ambientadores, esponjas, rin
farmacia y salud  → medicamentos, vitaminas, tiritas, termómetros, suplementos
tecnología        → electrónica, cables, pilas, bombillas inteligentes
electrodomésticos → aparatos con motor o calor (tostadora, aspiradora, batidora)
mobiliario        → muebles, estanterías, sillas, mesas, camas, sofás, almacenaje, colchón
textil y ropa     → ropa, calzado, ropa de cama, toallas, cortinas, cojines
papelería y oficina → papel, bolígrafos, carpetas, post-its, tijeras
otros             → lo que no encaje claramente (tuppers, velas decorativas…)

─── REGLA CRÍTICA: CAMPO TIENDA ───────────────────────────────────────────
El campo "tienda" es SIEMPRE obligatorio en cada item.
- Si el usuario menciona una tienda → úsala (en minúsculas)
- Si el usuario NO menciona tienda → usa exactamente la cadena "sin tienda"
- NUNCA dejes el campo "tienda" vacío o ausente

─── FORMATOS DE RESPUESTA ──────────────────────────────────────────────────
Añadir con tienda:
{"accion":"añadir","items":[{"producto":"leche","tienda":"mercadona","cantidad":"2","prioridad":"normal","categoria":"alimentación"}]}

Añadir sin tienda especificada:
{"accion":"añadir","items":[{"producto":"papel higiénico","tienda":"sin tienda","cantidad":"4","prioridad":"normal","categoria":"limpieza hogar"}]}

{"accion":"eliminar","productos":["leche","pan"]}
{"accion":"limpiar_tienda","tienda":"mercadona"}
{"accion":"limpiar_categoria","categoria":"mobiliario"}
{"accion":"actualizar","producto":"sofá","cambios":{"tienda":"sklum"}}
{"accion":"ver_categoria","categoria":"alimentación"}
{"accion":"ver_categorias","categorias":["alimentación","higiene personal"]}
{"accion":"ver_tienda","tienda":"mercadona"}
{"accion":"ver_todo"}
{"accion":"ver_urgentes"}
{"accion":"ver_filtro","tiendas":["mercadona","lidl"],"categorias":["alimentación"]}

NOTA sobre ver_filtro: úsalo cuando el usuario pida ver artículos filtrando por
varias tiendas Y/O varias categorías al mismo tiempo. Si sólo filtra por
una categoría usa ver_categoria; si sólo filtra por una tienda usa ver_tienda.
"""


def parsear_mensaje(mensaje: str) -> dict:
    """Envía el mensaje a Claude Haiku y devuelve el JSON parseado."""
    resp = claude.messages.create(
        model="claude-haiku-4-5",
        max_tokens=1000,
        system=_SYSTEM_PARSER,
        messages=[{"role": "user", "content": mensaje}],
    )
    texto = resp.content[0].text.strip()

    # Limpiar posibles backticks que el modelo ponga de más
    texto = texto.replace("```json", "").replace("```", "").strip()

    # Extraer el primer bloque JSON si hay texto extra
    inicio = texto.find("{")
    fin    = texto.rfind("}") + 1
    if inicio >= 0 and fin > inicio:
        texto = texto[inicio:fin]

    return json.loads(texto)


# ─────────────────────────────────────────
# BOTONES DE NAVEGACIÓN REUTILIZABLES
# ─────────────────────────────────────────

def _botones_navegacion() -> list[InlineKeyboardButton]:
    return [
        InlineKeyboardButton("📂 Categorías", callback_data="filtro_categorias"),
        InlineKeyboardButton("📍 Tiendas",    callback_data="filtro_tiendas"),
    ]


def _botones_ver_todo() -> list[InlineKeyboardButton]:
    return [InlineKeyboardButton("📋 Ver todo", callback_data="filtro_todo")]


# ─────────────────────────────────────────
# VISTAS
# ─────────────────────────────────────────

def formato_vista_completa(items: list[dict]) -> tuple[str, InlineKeyboardMarkup | None]:
    if not items:
        return "Tu lista está vacía.\n\n<i>Dime qué necesitas y lo añado.</i>", None

    por_categoria: dict[str, dict[str, list[dict]]] = defaultdict(lambda: defaultdict(list))
    for item in items:
        por_categoria[item["categoria"]][item["tienda"]].append(item)

    urgentes_total = sum(1 for i in items if i["prioridad"] == "urgente")
    lineas = [
        "🛒 <b>LISTA DE LA COMPRA</b>",
        f"<i>{len(items)} productos · {len(por_categoria)} categorías</i>",
        "",
    ]

    for categoria in sorted(por_categoria.keys()):
        lineas.append(f"{emoji_categoria(categoria)} <b>{categoria.upper()}</b>")
        for tienda in sorted(por_categoria[categoria].keys()):
            lineas.append(f"  📍 <i>{tienda.capitalize()}</i>")
            for p in por_categoria[categoria][tienda]:
                icon = "🔴" if p["prioridad"] == "urgente" else "⚪"
                bold = "<b>" if p["prioridad"] == "urgente" else ""
                endb = "</b>" if p["prioridad"] == "urgente" else ""
                lineas.append(f"    {icon} {bold}{p['producto']}{endb}  ×{p['cantidad']}")
        lineas.append("")

    if urgentes_total:
        lineas.append("<i>🔴 urgente · ⚪ normal</i>")

    botones = [
        _botones_navegacion(),
        [InlineKeyboardButton("🔴 Solo urgentes", callback_data="filtro_urgentes")],
    ]
    return "\n".join(lineas), InlineKeyboardMarkup(botones)


def formato_vista_por_categorias(items: list[dict]) -> tuple[str, InlineKeyboardMarkup | None]:
    if not items:
        return "La lista está vacía.", None

    por_cat: dict[str, list[dict]] = defaultdict(list)
    for item in items:
        por_cat[item["categoria"]].append(item)

    lineas = ["📂 <b>RESUMEN POR CATEGORÍAS</b>", ""]
    botones_cat = []
    for cat in sorted(por_cat.keys()):
        emoji  = emoji_categoria(cat)
        urgentes = sum(1 for p in por_cat[cat] if p["prioridad"] == "urgente")
        badge  = f"  🔴×{urgentes}" if urgentes else ""
        lineas.append(f"{emoji} <b>{cat.upper()}</b>{badge}  —  {len(por_cat[cat])} productos")
        botones_cat.append(
            InlineKeyboardButton(f"{emoji} {cat.capitalize()}", callback_data=f"categoria_{cat}")
        )

    lineas.append("\n<i>Pulsa una categoría para ver su detalle</i>")
    filas = [botones_cat[i:i + 2] for i in range(0, len(botones_cat), 2)]
    filas.append(_botones_ver_todo())
    return "\n".join(lineas), InlineKeyboardMarkup(filas)


def formato_vista_categoria(categoria: str, items: list[dict]) -> tuple[str, InlineKeyboardMarkup | None]:
    filtrados = [i for i in items if i["categoria"].lower() == categoria.lower()]
    if not filtrados:
        return f"No tienes nada en <b>{categoria}</b>.", None

    por_tienda: dict[str, list[dict]] = defaultdict(list)
    for item in filtrados:
        por_tienda[item["tienda"]].append(item)

    emoji = emoji_categoria(categoria)
    lineas = [f"{emoji} <b>{categoria.upper()}</b>", f"<i>{len(filtrados)} productos</i>", ""]
    for tienda in sorted(por_tienda.keys()):
        lineas.append(f"📍 <i>{tienda.capitalize()}</i>")
        for p in por_tienda[tienda]:
            icon = "🔴" if p["prioridad"] == "urgente" else "⚪"
            bold = "<b>" if p["prioridad"] == "urgente" else ""
            endb = "</b>" if p["prioridad"] == "urgente" else ""
            lineas.append(f"  {icon} {bold}{p['producto']}{endb}  ×{p['cantidad']}")
        lineas.append("")

    botones = [_botones_navegacion(), _botones_ver_todo()]
    return "\n".join(lineas), InlineKeyboardMarkup(botones)


def formato_vista_multicategoria(categorias: list[str], items: list[dict]) -> tuple[str, InlineKeyboardMarkup | None]:
    cats_lower = [c.lower() for c in categorias]
    filtrados  = [i for i in items if i["categoria"].lower() in cats_lower]
    if not filtrados:
        return "No tienes nada en esas categorías.", None

    por_cat: dict[str, dict[str, list[dict]]] = defaultdict(lambda: defaultdict(list))
    for item in filtrados:
        por_cat[item["categoria"]][item["tienda"]].append(item)

    nombres = " · ".join(f"{emoji_categoria(c)} {c}" for c in sorted(categorias))
    lineas  = [f"<b>{nombres.upper()}</b>", f"<i>{len(filtrados)} productos</i>", ""]

    for cat in sorted(por_cat.keys()):
        lineas.append(f"{emoji_categoria(cat)} <b>{cat.upper()}</b>")
        for tienda in sorted(por_cat[cat].keys()):
            lineas.append(f"  📍 <i>{tienda.capitalize()}</i>")
            for p in por_cat[cat][tienda]:
                icon = "🔴" if p["prioridad"] == "urgente" else "⚪"
                lineas.append(f"    {icon} {p['producto']}  ×{p['cantidad']}")
        lineas.append("")

    botones = [_botones_navegacion(), _botones_ver_todo()]
    return "\n".join(lineas), InlineKeyboardMarkup(botones)


def formato_vista_tiendas(items: list[dict]) -> tuple[str, InlineKeyboardMarkup | None]:
    if not items:
        return "La lista está vacía.", None

    por_tienda: dict[str, list[dict]] = defaultdict(list)
    for item in items:
        por_tienda[item["tienda"]].append(item)

    lineas = ["📍 <b>RESUMEN POR TIENDAS</b>", ""]
    botones_tienda = []
    for tienda in sorted(por_tienda.keys()):
        urgentes = sum(1 for p in por_tienda[tienda] if p["prioridad"] == "urgente")
        badge    = f"  🔴×{urgentes}" if urgentes else ""
        lineas.append(f"<b>{tienda.upper()}</b>{badge}  —  {len(por_tienda[tienda])} productos")
        botones_tienda.append(
            InlineKeyboardButton(f"📍 {tienda.capitalize()}", callback_data=f"tienda_{tienda.lower()}")
        )

    lineas.append("\n<i>Pulsa una tienda para ver su lista</i>")
    filas = [botones_tienda[i:i + 2] for i in range(0, len(botones_tienda), 2)]
    filas.append(_botones_ver_todo())
    return "\n".join(lineas), InlineKeyboardMarkup(filas)


def formato_vista_tienda(tienda: str, items: list[dict]) -> tuple[str, InlineKeyboardMarkup | None]:
    filtrados = [i for i in items if tienda.lower() in i["tienda"].lower()]
    if not filtrados:
        return f"No tienes nada pendiente en <b>{tienda.upper()}</b>.", None

    por_cat: dict[str, list[dict]] = defaultdict(list)
    for item in filtrados:
        por_cat[item["categoria"]].append(item)

    lineas = [
        f"📍 <b>{tienda.upper()}</b>",
        f"<i>{len(filtrados)} productos pendientes</i>",
        "",
    ]
    for cat in sorted(por_cat.keys()):
        lineas.append(f"{emoji_categoria(cat)} <i>{cat.capitalize()}</i>")
        for p in por_cat[cat]:
            icon = "🔴" if p["prioridad"] == "urgente" else "⚪"
            bold = "<b>" if p["prioridad"] == "urgente" else ""
            endb = "</b>" if p["prioridad"] == "urgente" else ""
            lineas.append(f"  {icon} {bold}{p['producto']}{endb}  ×{p['cantidad']}")
        lineas.append("")

    botones = [
        [InlineKeyboardButton("✅ Compra hecha — borrar tienda", callback_data=f"limpiar_{tienda.lower()}")],
        _botones_navegacion(),
        _botones_ver_todo(),
    ]
    return "\n".join(lineas), InlineKeyboardMarkup(botones)


def formato_vista_urgentes(items: list[dict]) -> tuple[str, InlineKeyboardMarkup | None]:
    urgentes = [i for i in items if i["prioridad"] == "urgente"]
    if not urgentes:
        return "No tienes nada urgente pendiente. 🎉", None

    por_tienda: dict[str, list[dict]] = defaultdict(list)
    for item in urgentes:
        por_tienda[item["tienda"]].append(item)

    lineas = ["🔴 <b>URGENTE</b>", f"<i>{len(urgentes)} productos</i>", ""]
    for tienda in sorted(por_tienda.keys()):
        lineas.append(f"📍 <b>{tienda.upper()}</b>")
        for p in por_tienda[tienda]:
            lineas.append(f"  <b>{p['producto']}</b>  ×{p['cantidad']}  <i>({p['categoria']})</i>")
        lineas.append("")

    botones = [_botones_ver_todo()]
    return "\n".join(lineas), InlineKeyboardMarkup(botones)


def formato_vista_filtro_combinado(
    tiendas: list[str],
    categorias: list[str],
    items: list[dict],
) -> tuple[str, InlineKeyboardMarkup | None]:
    """Filtra por tiendas Y/O categorías simultáneamente."""
    tiendas_lower    = [t.lower() for t in tiendas]
    categorias_lower = [c.lower() for c in categorias]

    def coincide(item: dict) -> bool:
        ok_tienda    = (not tiendas_lower)    or (item["tienda"].lower()    in tiendas_lower)
        ok_categoria = (not categorias_lower) or (item["categoria"].lower() in categorias_lower)
        return ok_tienda and ok_categoria

    filtrados = [i for i in items if coincide(i)]
    if not filtrados:
        return "No hay productos que coincidan con ese filtro.", None

    por_cat: dict[str, dict[str, list[dict]]] = defaultdict(lambda: defaultdict(list))
    for item in filtrados:
        por_cat[item["categoria"]][item["tienda"]].append(item)

    partes = []
    if tiendas:
        partes.append("📍 " + ", ".join(t.capitalize() for t in tiendas))
    if categorias:
        partes.append("📂 " + ", ".join(c.capitalize() for c in categorias))
    encabezado = " · ".join(partes) if partes else "Filtro"

    lineas = [f"<b>{encabezado.upper()}</b>", f"<i>{len(filtrados)} productos</i>", ""]
    for cat in sorted(por_cat.keys()):
        lineas.append(f"{emoji_categoria(cat)} <b>{cat.upper()}</b>")
        for tienda in sorted(por_cat[cat].keys()):
            lineas.append(f"  📍 <i>{tienda.capitalize()}</i>")
            for p in por_cat[cat][tienda]:
                icon = "🔴" if p["prioridad"] == "urgente" else "⚪"
                lineas.append(f"    {icon} {p['producto']}  ×{p['cantidad']}")
        lineas.append("")

    botones = [_botones_navegacion(), _botones_ver_todo()]
    return "\n".join(lineas), InlineKeyboardMarkup(botones)


# ─────────────────────────────────────────
# CONFIRMACIÓN DE BORRADO
# ─────────────────────────────────────────

def _preview_borrado(accion: str, parsed: dict, items: list[dict]) -> tuple[str, list[dict]]:
    """
    Devuelve (texto_preview, lista_de_items_afectados).
    El texto describe lo que se va a borrar.
    """
    if accion == "eliminar":
        productos = [p.lower() for p in parsed.get("productos", [])]
        afectados = [i for i in items if i["producto"].lower() in productos]
        if not afectados:
            nombres = ", ".join(parsed.get("productos", []))
            return f"No encontré <b>{nombres}</b> en la lista.", []
        lineas = ["🗑 <b>¿Eliminar estos productos?</b>", ""]
        for p in afectados:
            lineas.append(f"  ⚪ <b>{p['producto']}</b>  <i>({p['tienda']} · {p['categoria']})</i>")
        return "\n".join(lineas), afectados

    elif accion == "limpiar_tienda":
        tienda    = parsed.get("tienda", "").lower()
        afectados = [i for i in items if i["tienda"].lower() == tienda]
        if not afectados:
            return f"No hay productos de <b>{tienda.capitalize()}</b> en la lista.", []
        lineas = [
            f"🗑 <b>¿Borrar toda la compra de {tienda.capitalize()}?</b>",
            f"<i>{len(afectados)} productos se eliminarán:</i>",
            "",
        ]
        for p in afectados:
            lineas.append(f"  ⚪ {p['producto']}  ×{p['cantidad']}")
        return "\n".join(lineas), afectados

    elif accion == "limpiar_categoria":
        categoria = parsed.get("categoria", "").lower()
        afectados = [i for i in items if i["categoria"].lower() == categoria]
        if not afectados:
            return f"No hay productos en <b>{categoria}</b>.", []
        emoji = emoji_categoria(categoria)
        lineas = [
            f"🗑 <b>¿Borrar toda la categoría {emoji} {categoria}?</b>",
            f"<i>{len(afectados)} productos se eliminarán:</i>",
            "",
        ]
        for p in afectados:
            lineas.append(f"  ⚪ {p['producto']}  <i>({p['tienda']})</i>")
        return "\n".join(lineas), afectados

    return "¿Confirmas el borrado?", []


def necesita_confirmacion(afectados: list[dict]) -> bool:
    """Devuelve True si el borrado afecta a más de UMBRAL_CONFIRMACION artículos."""
    return len(afectados) > UMBRAL_CONFIRMACION


# ─────────────────────────────────────────
# AGENTE PRINCIPAL
# ─────────────────────────────────────────

async def agente_compra(mensaje: str) -> tuple[str, InlineKeyboardMarkup | None]:
    """
    Versión simple (sin gestión de confirmaciones).
    Usada internamente y para las acciones que no son de borrado.
    """
    try:
        parsed         = parsear_mensaje(mensaje)
        accion         = parsed.get("accion", "ver_todo")
        items_actuales = leer_items()

        return await _ejecutar_accion(accion, parsed, items_actuales)

    except json.JSONDecodeError:
        return (
            "No entendí el mensaje. Prueba: <i>añade leche y pan de Mercadona</i>",
            None,
        )
    except Exception as e:
        logger.error(f"agente_compra: {e}", exc_info=True)
        return "Hubo un error procesando tu lista. Inténtalo de nuevo.", None


async def _ejecutar_accion(
    accion: str,
    parsed: dict,
    items_actuales: list[dict],
) -> tuple[str, InlineKeyboardMarkup | None]:
    """Ejecuta cualquier acción (excepto las de borrado, que las gestiona agente_compra_con_confirmacion)."""

    if accion == "añadir":
        nuevos = parsed.get("items", [])
        if not nuevos:
            return "No entendí qué quieres añadir. Prueba: <i>leche x2 en Mercadona</i>", None

        existentes = {(i["producto"].lower(), i["tienda"].lower()) for i in items_actuales}
        añadidos: list[dict] = []

        for item in nuevos:
            # Sanear campos obligatorios
            item.setdefault("tienda",    "sin tienda")
            item.setdefault("cantidad",  "1")
            item.setdefault("prioridad", "normal")
            item.setdefault("categoria", "otros")
            item["tienda"]    = item["tienda"].lower().strip()
            item["categoria"] = item["categoria"].lower().strip()
            if item["categoria"] not in CATEGORIAS_VALIDAS:
                item["categoria"] = "otros"
            if item["prioridad"] not in ("urgente", "normal"):
                item["prioridad"] = "normal"

            key = (item["producto"].lower(), item["tienda"].lower())
            if key not in existentes:
                items_actuales.append(item)
                añadidos.append(item)
                existentes.add(key)

        if not añadidos:
            return "Esos productos ya estaban en la lista.", None

        # ── ESCRITURA ── sólo aquí, tras haber ampliado items_actuales
        guardar_excel(items_actuales)

        lineas = ["✅ <b>Añadido a la lista:</b>", ""]
        for i in añadidos:
            prior = "🔴" if i["prioridad"] == "urgente" else "⚪"
            lineas.append(
                f"  {prior} <b>{i['producto']}</b>  ×{i['cantidad']}\n"
                f"      {emoji_categoria(i['categoria'])} <i>{i['categoria']} · {i['tienda']}</i>"
            )
        botones = [[InlineKeyboardButton("📋 Ver lista completa", callback_data="filtro_todo")]]
        return "\n".join(lineas), InlineKeyboardMarkup(botones)

    elif accion == "actualizar":
        nombre_buscar = parsed.get("producto", "").lower()
        cambios       = parsed.get("cambios", {})
        encontrado    = None
        for item in items_actuales:
            if nombre_buscar in item["producto"].lower() or item["producto"].lower() in nombre_buscar:
                encontrado = item
                break
        if not encontrado:
            return (
                f"No encontré <b>{parsed.get('producto', '')}</b> en la lista.\n"
                "<i>¿Quizás está con otro nombre?</i>",
                None,
            )
        nombre_original = encontrado["producto"]
        cambios_aplicados = []
        if "tienda" in cambios:
            encontrado["tienda"] = cambios["tienda"].lower()
            cambios_aplicados.append(f"tienda → <i>{cambios['tienda']}</i>")
        if "cantidad" in cambios:
            encontrado["cantidad"] = cambios["cantidad"]
            cambios_aplicados.append(f"cantidad → <i>×{cambios['cantidad']}</i>")
        if "prioridad" in cambios:
            encontrado["prioridad"] = cambios["prioridad"]
            emoji_p = "🔴" if cambios["prioridad"] == "urgente" else "⚪"
            cambios_aplicados.append(f"prioridad → {emoji_p} <i>{cambios['prioridad']}</i>")
        if "categoria" in cambios:
            encontrado["categoria"] = cambios["categoria"].lower()
            cambios_aplicados.append(f"categoría → <i>{cambios['categoria']}</i>")

        guardar_excel(items_actuales)
        resumen = "\n  ".join(cambios_aplicados)
        botones = [[InlineKeyboardButton("📋 Ver lista", callback_data="filtro_todo")]]
        return f"✏️ <b>{nombre_original}</b> actualizado:\n  {resumen}", InlineKeyboardMarkup(botones)

    elif accion == "ver_urgentes":
        return formato_vista_urgentes(items_actuales)

    elif accion == "ver_tienda":
        return formato_vista_tienda(parsed.get("tienda", ""), items_actuales)

    elif accion == "ver_categoria":
        return formato_vista_categoria(parsed.get("categoria", "otros"), items_actuales)

    elif accion == "ver_categorias":
        return formato_vista_multicategoria(parsed.get("categorias", []), items_actuales)

    elif accion == "ver_filtro":
        tiendas    = parsed.get("tiendas",    [])
        categorias = parsed.get("categorias", [])
        return formato_vista_filtro_combinado(tiendas, categorias, items_actuales)

    else:  # ver_todo y cualquier cosa no reconocida
        return formato_vista_completa(items_actuales)


# ─────────────────────────────────────────
# AGENTE CON CONFIRMACIÓN DE BORRADO
# ─────────────────────────────────────────

async def agente_compra_con_confirmacion(mensaje: str) -> dict:
    """
    Punto de entrada principal desde main.py.
    - Si la acción es de borrado, devuelve un dict con tipo='confirmacion'
      cuando el número de artículos afectados supera UMBRAL_CONFIRMACION.
    - Si el borrado es pequeño (≤ UMBRAL_CONFIRMACION), lo ejecuta directamente.
    - Para el resto de acciones devuelve tipo='normal'.
    """
    try:
        parsed = parsear_mensaje(mensaje)
        accion = parsed.get("accion", "ver_todo")

        if accion in ACCIONES_QUE_BORRAN:
            items_actuales = leer_items()
            texto_preview, afectados = _preview_borrado(accion, parsed, items_actuales)

            # Sin artículos afectados → informar sin confirmar
            if not afectados:
                return {"tipo": "normal", "texto": texto_preview, "teclado": None}

            # Pocos artículos → ejecutar directamente sin pedir confirmación
            if not necesita_confirmacion(afectados):
                texto, teclado = await ejecutar_borrado_confirmado(accion, parsed)
                return {"tipo": "normal", "texto": texto, "teclado": teclado}

            # Muchos artículos → pedir confirmación
            return {
                "tipo":   "confirmacion",
                "accion": accion,
                "datos":  parsed,
                "texto":  texto_preview,
            }

        # Acción no destructiva
        texto, teclado = await _ejecutar_accion(accion, parsed, leer_items())
        return {"tipo": "normal", "texto": texto, "teclado": teclado}

    except Exception as e:
        logger.error(f"agente_compra_con_confirmacion: {e}", exc_info=True)
        return {"tipo": "normal", "texto": "Hubo un error. Inténtalo de nuevo.", "teclado": None}


async def ejecutar_borrado_confirmado(
    accion: str, parsed: dict
) -> tuple[str, InlineKeyboardMarkup | None]:
    """Ejecuta el borrado real (ya confirmado o sin necesidad de confirmación)."""
    items_actuales = leer_items()
    botones = [_botones_ver_todo()]

    if accion == "eliminar":
        productos_borrar = [p.lower() for p in parsed.get("productos", [])]
        antes            = len(items_actuales)
        items_actuales   = [i for i in items_actuales if i["producto"].lower() not in productos_borrar]
        borrados         = antes - len(items_actuales)
        guardar_excel(items_actuales)
        nombres = ", ".join(parsed.get("productos", []))
        return (
            f"✅ <b>Eliminado{' ('+str(borrados)+')' if borrados != 1 else ''}:</b> <i>{nombres}</i>",
            InlineKeyboardMarkup(botones),
        )

    elif accion == "limpiar_tienda":
        tienda         = parsed.get("tienda", "").lower()
        antes          = len(items_actuales)
        items_actuales = [i for i in items_actuales if i["tienda"].lower() != tienda]
        borrados       = antes - len(items_actuales)
        guardar_excel(items_actuales)
        return (
            f"✅ <b>Compra de {tienda.capitalize()} completada.</b>\n"
            f"<i>{borrados} productos eliminados.</i>",
            InlineKeyboardMarkup(botones),
        )

    elif accion == "limpiar_categoria":
        categoria      = parsed.get("categoria", "").lower()
        antes          = len(items_actuales)
        items_actuales = [i for i in items_actuales if i["categoria"].lower() != categoria]
        borrados       = antes - len(items_actuales)
        guardar_excel(items_actuales)
        return (
            f"✅ {emoji_categoria(categoria)} <b>{categoria.capitalize()} limpiado.</b>\n"
            f"<i>{borrados} productos eliminados.</i>",
            InlineKeyboardMarkup(botones),
        )

    return "Acción no reconocida.", None


# ─────────────────────────────────────────
# MANEJADOR DE CALLBACKS INLINE (navegación)
# ─────────────────────────────────────────

async def manejar_callback_compra(data: str) -> tuple[str, InlineKeyboardMarkup | None]:
    """
    Gestiona los callbacks de los botones inline de navegación.
    Los callbacks de borrado (limpiar_*) devuelven texto + None;
    main.py detecta eso y añade los botones de confirmación.
    """
    items_actuales = leer_items()

    if data == "filtro_todo":
        return formato_vista_completa(items_actuales)
    elif data == "filtro_urgentes":
        return formato_vista_urgentes(items_actuales)
    elif data == "filtro_tiendas":
        return formato_vista_tiendas(items_actuales)
    elif data == "filtro_categorias":
        return formato_vista_por_categorias(items_actuales)
    elif data.startswith("tienda_"):
        return formato_vista_tienda(data.removeprefix("tienda_"), items_actuales)
    elif data.startswith("categoria_"):
        return formato_vista_categoria(data.removeprefix("categoria_"), items_actuales)
    elif data.startswith("limpiar_"):
        # Genera el preview; main.py añadirá los botones de confirmar/cancelar
        tienda    = data.removeprefix("limpiar_")
        afectados = [i for i in items_actuales if i["tienda"].lower() == tienda]
        if not afectados:
            return f"No hay productos de <b>{tienda.capitalize()}</b>.", None
        lineas = [
            f"🗑 <b>¿Borrar toda la compra de {tienda.capitalize()}?</b>",
            f"<i>{len(afectados)} productos se eliminarán:</i>",
            "",
        ]
        for p in afectados:
            lineas.append(f"  ⚪ {p['producto']}  ×{p['cantidad']}")
        return "\n".join(lineas), None

    return "Acción no reconocida.", None