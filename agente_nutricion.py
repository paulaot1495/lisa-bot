"""
agente_nutricion.py — Agente de seguimiento nutricional para Lisa.
"""

import os
import re
import json
import logging
from datetime import datetime, timedelta
from pathlib import Path

from anthropic import Anthropic
from telegram import InlineKeyboardButton, InlineKeyboardMarkup

logger = logging.getLogger(__name__)
claude = Anthropic(api_key=os.getenv("ANTHROPIC_API_KEY"))
EXCEL_PATH = Path(os.getenv("NUTRITION_EXCEL_PATH", "/data/comidas.xlsx"))

PALABRAS_COMIDA = [
    "comí", "comi", "desayuné", "desayune", "almorcé", "almorce",
    "merendé", "merende", "cené", "cene", "tomé", "tome", "bebí", "bebi",
    "desayuno", "almuerzo", "comida", "merienda", "cena", "snack",
    "he comido", "he desayunado", "he cenado", "he almorzado",
    "me he comido", "me tomé", "ayer comí", "ayer cené", "ayer desayuné",
]

PALABRAS_RESET = [
    "borrar comidas", "borrar registro", "resetear comidas", "resetear nutricion",
    "borrar nutricion", "empezar de cero", "limpiar excel", "limpiar comidas",
    "eliminar comidas", "reset comidas",
]

def es_mensaje_nutricion(mensaje: str) -> bool:
    m = mensaje.lower()
    return any(p in m for p in PALABRAS_COMIDA)

def es_mensaje_reset_nutricion(mensaje: str) -> bool:
    m = mensaje.lower()
    return any(p in m for p in PALABRAS_RESET)

def detectar_fecha(mensaje: str) -> datetime:
    m = mensaje.lower()
    if any(p in m for p in ["ayer", "anoche", "ayer por"]):
        return datetime.now() - timedelta(days=1)
    return datetime.now()

def cargar_base_nutricional() -> dict:
    ruta = Path(os.getenv("BASE_NUTRICIONAL_PATH", "/data/base_nutricional.xlsx"))
    if not ruta.exists():
        return {}
    try:
        import pandas as pd
        df = pd.read_excel(ruta)
        df.columns = [c.strip().lower() for c in df.columns]
        base = {}
        for _, row in df.iterrows():
            nombre = str(row.get("alimento", row.get("nombre", ""))).strip().lower()
            if nombre:
                base[nombre] = {
                    "calorias":      float(row.get("calorias", row.get("kcal", 0)) or 0),
                    "proteinas":     float(row.get("proteinas", row.get("proteina", 0)) or 0),
                    "carbohidratos": float(row.get("carbohidratos", row.get("hidratos", 0)) or 0),
                    "grasas":        float(row.get("grasas", row.get("grasa", 0)) or 0),
                    "azucar":        float(row.get("azucar", 0) or 0),
                    "fibra":         float(row.get("fibra", 0) or 0),
                }
        return base
    except Exception as e:
        logger.warning(f"No se pudo cargar base_nutricional: {e}")
        return {}

SYSTEM_NUTRICION = """Eres un experto en nutricion. Analiza la comida descrita y devuelve SOLO un JSON valido, sin texto adicional, sin markdown.

Formato obligatorio:
{
  "alimentos": [
    {
      "nombre": "nombre del alimento",
      "cantidad_g": 150,
      "calorias": 200,
      "proteinas": 15,
      "carbohidratos": 20,
      "grasas": 5,
      "azucar": 3,
      "fibra": 2
    }
  ],
  "totales": {
    "calorias": 200,
    "proteinas": 15,
    "carbohidratos": 20,
    "grasas": 5,
    "azucar": 3,
    "fibra": 2
  },
  "descripcion_comida": "Breve descripcion de lo que se ha comido"
}

Reglas:
- Todos los valores son por la cantidad descrita (no por 100g)
- Si no se especifica cantidad, usa una porcion estandar razonable
- Usa la base nutricional proporcionada cuando el alimento aparezca; si no, usa tus conocimientos
- Redondea a 1 decimal
- calorias en kcal, todo lo demas en gramos"""

async def analizar_nutricion(mensaje: str, base_nutricional: dict) -> dict:
    base_str = ""
    if base_nutricional:
        base_str = f"\n\nBASE NUTRICIONAL DISPONIBLE (por 100g):\n{json.dumps(base_nutricional, ensure_ascii=False, indent=2)}"
    prompt = f"Analiza esta comida y extrae los valores nutricionales:{base_str}\n\nMensaje del usuario: {mensaje}"
    resp = claude.messages.create(
        model="claude-haiku-4-5",
        max_tokens=1024,
        system=SYSTEM_NUTRICION,
        messages=[{"role": "user", "content": prompt}]
    )
    raw = resp.content[0].text.strip()
    raw = re.sub(r"```json\s*|\s*```", "", raw).strip()
    return json.loads(raw)

COLUMNAS = [
    "Fecha", "Descripcion",
    "Calorias (kcal)", "Proteinas (g)", "Carbohidratos (g)",
    "Grasas (g)", "Azucar (g)", "Fibra (g)"
]
COL_IDX      = {c: i+1 for i, c in enumerate(COLUMNAS)}
C_HEADER_BG  = "6B8CAE"
C_HEADER_FG  = "FFFFFF"
C_ROW1_BG    = "EAF4FB"
C_ROW2_BG    = "F5FBFE"
BORDER_COLOR = "BBCDD8"
NUM_FMT      = "#,##0.0"

def _get_styles():
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    thin   = Side(style='thin', color=BORDER_COLOR)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    def apply(cell, bold=False, fg="000000", bg=None, align="center", size=10, italic=False):
        cell.font      = Font(name="Arial", bold=bold, size=size, color=fg, italic=italic)
        if bg:
            cell.fill  = PatternFill("solid", start_color=bg)
        cell.alignment = Alignment(horizontal=align, vertical="center",
                                   indent=1 if align == "left" else 0)
        cell.border    = border
    return apply

def crear_excel_nuevo():
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Comidas"
    apply = _get_styles()
    ws.merge_cells(f"A1:{chr(64+len(COLUMNAS))}1")
    c = ws["A1"]
    c.value = "Registro Nutricional"
    apply(c, bold=True, fg=C_HEADER_FG, bg=C_HEADER_BG, size=13)
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 6
    for i, col in enumerate(COLUMNAS, 1):
        cell = ws.cell(row=3, column=i, value=col)
        apply(cell, bold=True, fg=C_HEADER_FG, bg=C_HEADER_BG, size=11)
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 40
    for col_letter in "CDEFGH":
        ws.column_dimensions[col_letter].width = 18
    EXCEL_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(EXCEL_PATH)
    return wb

def cargar_o_crear_excel():
    from openpyxl import load_workbook
    if not EXCEL_PATH.exists():
        return crear_excel_nuevo(), True
    wb = load_workbook(EXCEL_PATH)
    return wb, False

def resetear_excel() -> bool:
    try:
        if EXCEL_PATH.exists():
            EXCEL_PATH.unlink()
        crear_excel_nuevo()
        return True
    except Exception as e:
        logger.error(f"Error reseteando Excel: {e}")
        return False

def fila_para_fecha(ws, fecha_str: str):
    for row in ws.iter_rows(min_row=4, max_col=1):
        val = row[0].value
        if val and str(val).strip() == fecha_str:
            return row[0].row
    return None

def siguiente_fila_disponible(ws) -> int:
    max_row = 3
    for row in ws.iter_rows(min_row=4, max_col=1):
        if row[0].value is not None:
            max_row = row[0].row
    return max(4, max_row + 1)

def escribir_fila(ws, row_num, fecha_str, datos, es_nueva_fila):
    apply = _get_styles()
    bg = C_ROW1_BG if row_num % 2 == 0 else C_ROW2_BG

    def set_cell(col_name, value, fmt=None, align="center"):
        col  = COL_IDX[col_name]
        cell = ws.cell(row=row_num, column=col, value=value)
        apply(cell, fg="000000", bg=bg, align=align)
        if fmt:
            cell.number_format = fmt

    if es_nueva_fila:
        set_cell("Fecha",             fecha_str,                          align="center")
        set_cell("Descripcion",       datos["descripcion_comida"],        align="left")
        set_cell("Calorias (kcal)",   datos["totales"]["calorias"],       NUM_FMT)
        set_cell("Proteinas (g)",     datos["totales"]["proteinas"],      NUM_FMT)
        set_cell("Carbohidratos (g)", datos["totales"]["carbohidratos"],  NUM_FMT)
        set_cell("Grasas (g)",        datos["totales"]["grasas"],         NUM_FMT)
        set_cell("Azucar (g)",        datos["totales"]["azucar"],         NUM_FMT)
        set_cell("Fibra (g)",         datos["totales"]["fibra"],          NUM_FMT)
    else:
        campos = [
            ("Calorias (kcal)",   "calorias"),
            ("Proteinas (g)",     "proteinas"),
            ("Carbohidratos (g)", "carbohidratos"),
            ("Grasas (g)",        "grasas"),
            ("Azucar (g)",        "azucar"),
            ("Fibra (g)",         "fibra"),
        ]
        for col_name, key in campos:
            col  = COL_IDX[col_name]
            cell = ws.cell(row=row_num, column=col)
            actual = float(cell.value or 0)
            cell.value = round(actual + datos["totales"][key], 1)
            apply(cell, fg="000000", bg=bg)
            cell.number_format = NUM_FMT
        col_desc  = COL_IDX["Descripcion"]
        cell_desc = ws.cell(row=row_num, column=col_desc)
        desc_actual = str(cell_desc.value or "")
        cell_desc.value = desc_actual + " + " + datos["descripcion_comida"]
        apply(cell_desc, fg="000000", bg=bg, align="left")

    ws.row_dimensions[row_num].height = 20

def actualizar_excel(datos, fecha):
    try:
        wb, _ = cargar_o_crear_excel()
        ws = wb["Comidas"]
        fecha_str = fecha.strftime("%d/%m/%Y")
        fila_existente = fila_para_fecha(ws, fecha_str)
        es_nueva = fila_existente is None
        row_num  = siguiente_fila_disponible(ws) if es_nueva else fila_existente
        escribir_fila(ws, row_num, fecha_str, datos, es_nueva)
        wb.save(EXCEL_PATH)
        return True, es_nueva
    except Exception as e:
        logger.error(f"Error actualizando Excel: {e}")
        return False, False

def construir_respuesta(datos, fecha_str, es_ayer, es_nueva_fila):
    t      = datos["totales"]
    accion = "Anadido al dia anterior" if es_ayer else "Registrado"
    estado = "nueva entrada" if es_nueva_fila else "sumado a lo anterior"

    lineas = ""
    for a in datos.get("alimentos", []):
        nombre   = a.get("nombre", "?")
        cantidad = a.get("cantidad_g", "?")
        kcal     = a.get("calorias", 0)
        prot     = a.get("proteinas", 0)
        ch       = a.get("carbohidratos", 0)
        grasas   = a.get("grasas", 0)
        lineas  += (
            f"  <b>{nombre}</b> ({cantidad}g)\n"
            f"    {kcal:.0f} kcal  |  {prot:.1f}g prot  |  {ch:.1f}g CH  |  {grasas:.1f}g grasa\n"
        )

    return (
        f"{'Dia anterior' if es_ayer else 'Hoy'} — <b>{fecha_str}</b> <i>({estado})</i>\n\n"
        f"<b>{datos['descripcion_comida']}</b>\n\n"
        f"<b>Desglose por alimento:</b>\n"
        f"{lineas}\n"
        f"<b>Total acumulado del dia:</b>\n"
        f"  Calorias:      <b>{t['calorias']:.0f} kcal</b>\n"
        f"  Proteinas:     <b>{t['proteinas']:.1f} g</b>\n"
        f"  Carbohidratos: <b>{t['carbohidratos']:.1f} g</b>\n"
        f"  Grasas:        <b>{t['grasas']:.1f} g</b>\n"
        f"  Azucar:        <b>{t['azucar']:.1f} g</b>\n"
        f"  Fibra:         <b>{t['fibra']:.1f} g</b>\n\n"
        f"<i>Excel actualizado.</i>"
    )

def borrar_fila_fecha(fecha_str: str) -> bool:
    """Elimina la fila del Excel correspondiente a fecha_str. Devuelve True si existia y se borro."""
    try:
        wb, _ = cargar_o_crear_excel()
        ws = wb["Comidas"]
        fila = fila_para_fecha(ws, fecha_str)
        if fila is None:
            return False
        ws.delete_rows(fila)
        wb.save(EXCEL_PATH)
        return True
    except Exception as e:
        logger.error(f"Error borrando fila {fecha_str}: {e}")
        return False

def teclado_confirmar_reset(fecha_str: str, scope: str) -> InlineKeyboardMarkup:
    """
    scope = "dia"  → borra solo esa fecha
    scope = "todo" → borra el Excel completo
    """
    fecha_enc = fecha_str.replace("/", "-")  # evita caracteres raros en callback_data
    return InlineKeyboardMarkup([[
        InlineKeyboardButton(
            "Si, borrar" ,
            callback_data=f"nutricion_reset_confirm_{scope}_{fecha_enc}"
        ),
        InlineKeyboardButton("Cancelar", callback_data="nutricion_reset_cancel"),
    ]])

async def agente_nutricion(mensaje: str):
    fecha     = detectar_fecha(mensaje)
    fecha_str = fecha.strftime("%d/%m/%Y")
    es_ayer   = (datetime.now() - fecha).days >= 1
    base      = cargar_base_nutricional()

    try:
        datos = await analizar_nutricion(mensaje, base)
    except Exception as e:
        logger.error(f"Error analisis: {e}")
        return "No pude analizar la informacion nutricional. Intentalo de nuevo.", None

    ok, es_nueva_fila = actualizar_excel(datos, fecha)
    if not ok:
        return "Error al guardar en el Excel. Revisa que /data este montado correctamente.", None

    return construir_respuesta(datos, fecha_str, es_ayer, es_nueva_fila), None

async def agente_nutricion_reset(mensaje: str):
    """
    Detecta si el usuario quiere borrar un dia concreto (hoy/ayer) o todo el registro.
    Muestra confirmacion con botones antes de actuar.
    """
    m = mensaje.lower()

    # ¿Quiere borrar todo?
    if any(p in m for p in ["todo", "registro", "excel", "cero", "completo"]):
        hoy = datetime.now().strftime("%d/%m/%Y")
        texto = (
            "<b>Seguro que quieres borrar TODO el registro de comidas?</b>\n\n"
            "Se eliminara el Excel completo y se creara uno nuevo vacio.\n"
            "<i>Esta accion no se puede deshacer.</i>"
        )
        return texto, teclado_confirmar_reset(hoy, "todo")

    # Por defecto: borrar solo el dia detectado (hoy o ayer)
    fecha     = detectar_fecha(mensaje)
    fecha_str = fecha.strftime("%d/%m/%Y")
    es_ayer   = (datetime.now() - fecha).days >= 1
    dia_label = "ayer" if es_ayer else "hoy"

    texto = (
        f"<b>Seguro que quieres borrar el registro de {dia_label} ({fecha_str})?</b>\n\n"
        f"Se eliminara solo la fila de ese dia del Excel.\n"
        f"<i>Esta accion no se puede deshacer.</i>"
    )
    return texto, teclado_confirmar_reset(fecha_str, "dia")

async def manejar_callback_nutricion(data: str):
    if data == "nutricion_reset_cancel":
        return "Cancelado. El registro sigue intacto.", None

    if data.startswith("nutricion_reset_confirm_"):
        # formato: nutricion_reset_confirm_{scope}_{fecha_enc}
        partes    = data.split("_", 4)   # ['nutricion','reset','confirm', scope, fecha_enc]
        scope     = partes[3]
        fecha_enc = partes[4] if len(partes) > 4 else ""
        fecha_str = fecha_enc.replace("-", "/")

        if scope == "todo":
            ok = resetear_excel()
            if ok:
                return "<b>Registro completo borrado.</b> El Excel esta vacio y listo para empezar de cero.", None
            return "No pude borrar el Excel. Intentalo de nuevo.", None

        if scope == "dia":
            ok = borrar_fila_fecha(fecha_str)
            if ok:
                return f"<b>Registro de {fecha_str} eliminado.</b> El resto del historial sigue intacto.", None
            return f"No encontre datos para {fecha_str}. Puede que ya estuviera vacio.", None

    return "Accion no reconocida.", None

# ──────────────────────────────────────────────
# CONSULTAS / ANÁLISIS DEL REGISTRO
# ──────────────────────────────────────────────

PALABRAS_CONSULTA = [
    "macros de hoy", "macros de ayer", "macros del dia", "que llevo hoy",
    "que llevo comido", "cuanto llevo", "resumen de hoy", "resumen de ayer",
    "resumen del dia", "resumen de la semana", "resumen semanal", "resumen del mes",
    "resumen mensual", "como voy", "como he ido", "tendencias", "analisis",
    "cuanto he comido", "cuantas calorias llevo", "cuantas calorias he comido",
    "mis macros", "mi registro", "ver mis datos", "ver registro", "historial",
    "media semanal", "media de la semana", "cuando como mas", "cuando como menos",
    "dia que mas", "dia que menos", "estadisticas",
]

def es_mensaje_consulta_nutricion(mensaje: str) -> bool:
    m = mensaje.lower()
    return any(p in m for p in PALABRAS_CONSULTA)

def leer_datos_excel(dias: int = 7) -> list[dict]:
    """
    Lee las ultimas N filas del Excel y las devuelve como lista de dicts.
    dias=1 → solo hoy, dias=7 → semana, dias=30 → mes, dias=0 → todo.
    """
    if not EXCEL_PATH.exists():
        return []
    try:
        import pandas as pd
        df = pd.read_excel(EXCEL_PATH, sheet_name="Comidas", header=2)
        df.columns = [str(c).strip() for c in df.columns]
        df = df.dropna(subset=[df.columns[0]])  # quitar filas sin fecha

        registros = []
        for _, row in df.iterrows():
            fecha_val = str(row.iloc[0]).strip()
            try:
                fecha_dt = datetime.strptime(fecha_val, "%d/%m/%Y")
            except ValueError:
                continue

            if dias > 0:
                limite = datetime.now() - timedelta(days=dias)
                if fecha_dt < limite:
                    continue

            registros.append({
                "fecha":          fecha_val,
                "descripcion":    str(row.iloc[1] if len(row) > 1 else ""),
                "calorias":       float(row.iloc[2] if len(row) > 2 else 0) or 0,
                "proteinas":      float(row.iloc[3] if len(row) > 3 else 0) or 0,
                "carbohidratos":  float(row.iloc[4] if len(row) > 4 else 0) or 0,
                "grasas":         float(row.iloc[5] if len(row) > 5 else 0) or 0,
                "azucar":         float(row.iloc[6] if len(row) > 6 else 0) or 0,
                "fibra":          float(row.iloc[7] if len(row) > 7 else 0) or 0,
            })
        return sorted(registros, key=lambda x: datetime.strptime(x["fecha"], "%d/%m/%Y"))
    except Exception as e:
        logger.error(f"Error leyendo Excel para consulta: {e}")
        return []

def detectar_rango_consulta(mensaje: str) -> int:
    """Devuelve el numero de dias a consultar segun el mensaje."""
    m = mensaje.lower()
    if any(p in m for p in ["mes", "mensual", "30 dias", "ultimo mes"]):
        return 30
    if any(p in m for p in ["semana", "semanal", "7 dias", "esta semana"]):
        return 7
    if any(p in m for p in ["ayer"]):
        return 2   # incluye ayer y hoy
    if any(p in m for p in ["hoy", "dia", "llevo", "voy"]):
        return 1
    return 7  # por defecto: semana

SYSTEM_ANALISIS = """Eres Lisa, una AI Manager personal especializada en nutricion.
Recibes datos del registro nutricional del usuario y debes analizarlos de forma clara, directa y util.

FORMATO OBLIGATORIO (nunca uses asteriscos ni guiones bajos):
- Para negrita usa: <b>texto</b>
- Para cursiva usa: <i>texto</i>
- Para listas usa guiones simples: - item
- Respuestas concisas y accionables

Analiza los datos y proporciona:
1. Un resumen claro de los numeros (medias, totales segun corresponda)
2. Patrones o tendencias que detectes (dias con mas calorias, variaciones, etc.)
3. Una observacion practica y util (sin ser condescendiente)

Si solo hay datos de un dia, haz el resumen de ese dia.
Si hay varios dias, incluye medias, el dia con mas y menos calorias, y tendencias.
Si no hay datos, indicalo claramente."""

async def agente_consulta_nutricion(mensaje: str) -> str:
    """Lee el Excel y pide a Claude que analice y responda la consulta del usuario."""
    dias      = detectar_rango_consulta(mensaje)
    registros = leer_datos_excel(dias)

    if not registros:
        periodo = {1: "hoy", 7: "esta semana", 30: "este mes"}.get(dias, f"los ultimos {dias} dias")
        return f"No tengo datos registrados para {periodo}. Cuéntame qué has comido y lo apunto."

    datos_str = json.dumps(registros, ensure_ascii=False, indent=2)
    hoy_str   = datetime.now().strftime("%d/%m/%Y")

    prompt = (
        f"Fecha de hoy: {hoy_str}\n"
        f"El usuario pregunta: \"{mensaje}\"\n\n"
        f"Datos del registro nutricional (ultimos {dias} dias):\n{datos_str}\n\n"
        f"Analiza estos datos y responde a la pregunta del usuario."
    )

    resp = claude.messages.create(
        model="claude-haiku-4-5",
        max_tokens=800,
        system=SYSTEM_ANALISIS,
        messages=[{"role": "user", "content": prompt}]
    )
    return resp.content[0].text.strip()